import os
import numpy as np
import pandas as pd
import tensorflow as tf
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing import image
from manager_models import ModelManager
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from collections import defaultdict

# Define the labels directory
LABEL_DIR = r"C:\Users\roman\PycharmProjects\pythonProject3\labels"

# Load the Best Model
model_path = ModelManager.get_best_model()

if model_path and os.path.exists(model_path):
    print(f"Loading model from: {model_path}")
    model = load_model(model_path)
    print("Model successfully loaded.\n")
else:
    raise FileNotFoundError(f"ERROR: Model not found at {model_path}. Please train the model first.")

# Get class names directly from directory structure
class_names = sorted([folder for folder in os.listdir(LABEL_DIR) if os.path.isdir(os.path.join(LABEL_DIR, folder))])
print(f"Using class names from directories: {class_names}\n")


# Function to classify a single image
def classify_image(img_path):
    img = image.load_img(img_path, target_size=(224, 224))
    img_array = image.img_to_array(img)
    img_array = np.expand_dims(img_array, axis=0) / 255.0  # Normalize

    predictions = model.predict(img_array)
    top_index = np.argmax(predictions)
    brand = class_names[top_index]
    confidence = predictions[0][top_index] * 100

    return brand, confidence


# Iterate through each sub-directory and classify images
results = []
class_stats = defaultdict(lambda: {"positive": 0, "negative": 0, "total": 0})

for sub_dir in class_names:
    sub_dir_path = os.path.join(LABEL_DIR, sub_dir)

    if os.path.isdir(sub_dir_path):
        image_files = [f for f in os.listdir(sub_dir_path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]

        for file in image_files:
            img_path = os.path.join(sub_dir_path, file)
            predicted_brand, confidence = classify_image(img_path)

            is_match = predicted_brand == sub_dir
            match_status = "Positive" if is_match else "Negative"

            # Update statistics for the actual class
            if is_match:
                class_stats[sub_dir]["positive"] += 1
            else:
                class_stats[sub_dir]["negative"] += 1
            class_stats[sub_dir]["total"] += 1

            results.append({
                "Image": file,
                "Actual": sub_dir,
                "Predicted": predicted_brand,
                "Confidence": f"{confidence:.2f}%",
                "Match": match_status
            })

            # Print results to terminal
            print(f"{file} | Actual: {sub_dir} | Predicted: {predicted_brand} ({confidence:.2f}%) | {match_status}")

# Calculate class accuracy percentages
for result in results:
    actual_class = result["Actual"]
    stats = class_stats[actual_class]
    if stats["total"] > 0:
        positive_percentage = (stats["positive"] / stats["total"]) * 100
        result["Class Accuracy"] = f"{positive_percentage:.2f}% ({stats['positive']}/{stats['total']})"
    else:
        result["Class Accuracy"] = "N/A"

print("\nClassification complete!")

# Calculate overall accuracy
total_images = len(results)
correct_predictions = sum(1 for r in results if "Positive" in r["Match"])
overall_accuracy = (correct_predictions / total_images * 100) if total_images > 0 else 0
print(f"Overall Accuracy: {overall_accuracy:.2f}% ({correct_predictions}/{total_images})")

# Convert results to a DataFrame
results_df = pd.DataFrame(results)

# Create a timestamp-based sheet name
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
sheet_name = f"Run_{timestamp}"

# Output path for Excel file
output_excel_path = os.path.join(os.path.dirname(LABEL_DIR), "classification_results.xlsx")

# Add summary row at the bottom
summary_row = pd.DataFrame([{
    "Image": "SUMMARY",
    "Actual": "",
    "Predicted": "",
    "Confidence": "",
    "Match": f"Overall: {overall_accuracy:.2f}%",
    "Class Accuracy": f"{correct_predictions}/{total_images}"
}])

# Append summary row to results
results_df = pd.concat([results_df, summary_row], ignore_index=True)

# Check if the Excel file already exists
if os.path.exists(output_excel_path):
    # For existing Excel files, use a different approach
    try:
        # Load existing workbook to get sheet names
        book = load_workbook(output_excel_path)
        existing_sheets = book.sheetnames

        # Check if sheet name already exists
        if sheet_name in existing_sheets:
            sheet_name = f"{sheet_name}_duplicate"

        # Use ExcelWriter with the 'with' context to safely handle file operations
        with pd.ExcelWriter(
                output_excel_path,
                engine='openpyxl',
                mode='a',
                if_sheet_exists='replace'
        ) as writer:
            # Write the results to the new sheet
            results_df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Results added to existing file as new sheet '{sheet_name}' in: {output_excel_path}")
    except Exception as e:
        print(f"Error adding sheet to existing file: {e}")
        print("Creating new file instead...")
        results_df.to_excel(output_excel_path, sheet_name=sheet_name, index=False)
else:
    # Create a new Excel file with the first sheet
    results_df.to_excel(output_excel_path, sheet_name=sheet_name, index=False)
    print(f"Results saved to new file as sheet '{sheet_name}' in: {output_excel_path}")

# Also save the latest results as CSV for backward compatibility (optional)
# output_csv_path = os.path.join(os.path.dirname(LABEL_DIR), "classification_results.csv")
# results_df.to_csv(output_csv_path, index=False)
# print(f"Latest results also saved to: {output_csv_path}")
